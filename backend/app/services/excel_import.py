"""Excel import (product-spec §7, integration brief EXCEL IMPORT).

`ExcelImportService.parse()` turns raw `.xlsx` bytes into unscheduled `Task`
domain objects (UUIDs assigned, predecessor names resolved to UUIDs) plus a
list of non-fatal warnings. It never touches the scheduler or the store.

Row-level errors are collected across the *whole* sheet before raising —
one `ImportValidationError` with every `{row, field, code, message}` issue,
not just the first.
"""

from __future__ import annotations

import io
import re
import uuid
from datetime import date, datetime

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.constants import (
    MAX_ASSIGNEE_LENGTH,
    MAX_DESCRIPTION_LENGTH,
    MAX_DURATION_WORKDAYS,
    MAX_FILE_SIZE_BYTES,
    MAX_TASK_NAME_LENGTH,
    MAX_TASKS,
    MIN_DURATION_WORKDAYS,
)
from app.domain.errors import (
    FileTooLargeError,
    ImportValidationError,
    InvalidWorkbookError,
    TooManyTasksError,
    UnsupportedFileTypeError,
)
from app.domain.models import Task
from app.domain.normalize import normalize_name

REQUIRED_COLUMNS = ("Задача", "Описание", "Исполнитель", "Длительность", "Предшественники")
OPTIONAL_COLUMNS = ("Плановая трудоёмкость, ч", "Дата начала", "Дата окончания", "Не ранее")

_PLACEHOLDER_DATE = date(2000, 1, 1)


class ParsedImport:
    def __init__(self, project_name: str, tasks: list[Task], warnings: list[str]) -> None:
        self.project_name = project_name
        self.tasks = tasks
        self.warnings = warnings


class ExcelImportService:
    def parse(self, content: bytes, *, filename: str) -> ParsedImport:
        _validate_extension(filename)
        _validate_size(content)
        project_name = _safe_project_name(filename)

        workbook = _open_workbook(content)
        try:
            sheet = _first_non_empty_sheet(workbook)
            header_row_index, columns = _read_header(sheet)
            self._check_merged_cells(sheet, header_row_index)
            return self._parse_rows(sheet, header_row_index, columns, project_name)
        finally:
            workbook.close()

    # -- worksheet-level checks -------------------------------------------------

    def _check_merged_cells(self, sheet: Worksheet, header_row_index: int) -> None:
        if not sheet.merged_cells.ranges:
            return
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row >= header_row_index:
                raise ImportValidationError(
                    "Объединённые ячейки не поддерживаются в области таблицы.",
                    details=[
                        {
                            "row": merged_range.min_row,
                            "field": None,
                            "code": "IMPORT_MERGED_CELLS_NOT_SUPPORTED",
                            "message": f"Объединённая ячейка {merged_range.coord} внутри таблицы.",
                        }
                    ],
                )

    # -- row parsing --------------------------------------------------------------

    def _parse_rows(
        self,
        sheet: Worksheet,
        header_row_index: int,
        columns: dict[str, int],
        project_name: str,
    ) -> ParsedImport:
        errors: list[dict] = []
        warnings: list[str] = []

        # pass 1: read + validate every non-empty row's own fields.
        rows: list[dict] = []
        name_to_row: dict[str, int] = {}

        for row_index in range(header_row_index + 1, sheet.max_row + 1):
            values = {
                col: sheet.cell(row=row_index, column=idx) for col, idx in columns.items()
            }
            if _row_is_empty(values):
                continue

            row_errors: list[dict] = []

            raw_name = values["Задача"]
            name = _parse_name(raw_name, row_index, row_errors)
            if name is not None:
                normalized = normalize_name(name)
                if normalized in name_to_row:
                    row_errors.append(
                        _issue(
                            row_index,
                            "Задача",
                            "IMPORT_DUPLICATE_TASK_NAME",
                            f"Название «{name}» повторяет строку {name_to_row[normalized]}.",
                        )
                    )
                else:
                    name_to_row[normalized] = row_index

            description = _parse_description(values["Описание"], row_index, row_errors)
            assignee = _parse_assignee(values["Исполнитель"], row_index, row_errors)
            duration = _parse_duration(values["Длительность"], row_index, row_errors)
            predecessor_names = _parse_predecessors(values["Предшественники"], row_index, row_errors)
            start_not_before = None
            if "Не ранее" in values:
                start_not_before = _parse_optional_date(
                    values["Не ранее"], row_index, "Не ранее", row_errors
                )

            errors.extend(row_errors)
            rows.append(
                {
                    "row": row_index,
                    "name": name,
                    "description": description,
                    "assignee": assignee,
                    "duration": duration,
                    "predecessor_names": predecessor_names,
                    "start_not_before": start_not_before,
                }
            )

        if len(rows) > MAX_TASKS:
            raise TooManyTasksError(
                f"Файл содержит {len(rows)} задач, максимум {MAX_TASKS}.",
                details=[{"row": None, "field": None, "code": "IMPORT_TOO_MANY_TASKS", "message": "too many tasks"}],
            )

        # pass 2: resolve predecessor names -> ids (forward references allowed).
        row_id: dict[int, uuid.UUID] = {r["row"]: uuid.uuid4() for r in rows}
        tasks: list[Task] = []

        for display_order, r in enumerate(rows, start=1):
            resolved_ids: list[uuid.UUID] = []
            seen: set[uuid.UUID] = set()
            for pred_name in r["predecessor_names"]:
                pred_normalized = normalize_name(pred_name)
                pred_row = name_to_row.get(pred_normalized)
                if pred_row is None:
                    errors.append(
                        _issue(
                            r["row"],
                            "Предшественники",
                            "IMPORT_UNKNOWN_PREDECESSOR",
                            f"Задача «{pred_name}» не найдена в файле.",
                        )
                    )
                    continue
                pred_id = row_id[pred_row]
                if pred_id == row_id[r["row"]]:
                    errors.append(
                        _issue(
                            r["row"],
                            "Предшественники",
                            "IMPORT_SELF_DEPENDENCY",
                            "Задача не может быть предшественником самой себя.",
                        )
                    )
                    continue
                if pred_id in seen:
                    warnings.append(
                        f"Строка {r['row']}: повторяющийся предшественник «{pred_name}» проигнорирован."
                    )
                    continue
                seen.add(pred_id)
                resolved_ids.append(pred_id)

            if r["name"] is None or r["duration"] is None:
                # Already recorded as an error above; skip building a Task for this row.
                continue

            tasks.append(
                Task(
                    id=row_id[r["row"]],
                    name=r["name"],
                    description=r["description"] or "",
                    assignee=r["assignee"],
                    duration_workdays=r["duration"],
                    predecessor_ids=resolved_ids,
                    start_not_before=r["start_not_before"],
                    start_date=_PLACEHOLDER_DATE,
                    end_date=_PLACEHOLDER_DATE,
                    display_order=display_order,
                    created_source="import",
                )
            )

        if errors:
            raise ImportValidationError(
                f"Импорт остановлен: найдено {len(errors)} ошибок.",
                details=errors,
            )

        if not tasks:
            raise ImportValidationError(
                "В файле нет ни одной задачи.",
                details=[{"row": None, "field": None, "code": "IMPORT_EMPTY", "message": "no tasks found"}],
            )

        return ParsedImport(project_name=project_name, tasks=tasks, warnings=warnings)


# --------------------------------------------------------------------------------
# module-level helpers
# --------------------------------------------------------------------------------


def _validate_extension(filename: str) -> None:
    if not filename.lower().endswith(".xlsx"):
        raise UnsupportedFileTypeError(f"Файл «{filename}» должен иметь расширение .xlsx.")


def _validate_size(content: bytes) -> None:
    if len(content) > MAX_FILE_SIZE_BYTES:
        raise FileTooLargeError(
            f"Файл превышает лимит {MAX_FILE_SIZE_BYTES // (1024 * 1024)} МБ."
        )


def _safe_project_name(filename: str) -> str:
    # Never treat the filename as a filesystem path — strip any directory
    # components a hostile client might smuggle into the multipart field.
    base = filename.replace("\\", "/").rsplit("/", 1)[-1]
    stem = base.rsplit(".", 1)[0] if "." in base else base
    stem = re.sub(r"[^\w\-. ]+", "_", stem, flags=re.UNICODE).strip()
    return (stem or "project")[:200]


def _open_workbook(content: bytes):
    try:
        return openpyxl.load_workbook(io.BytesIO(content), data_only=False, read_only=False)
    except Exception as exc:  # noqa: BLE001 - any parse failure means "not a valid workbook"
        raise InvalidWorkbookError(f"Не удалось открыть файл как Excel-книгу: {exc}") from exc


def _first_non_empty_sheet(workbook) -> Worksheet:
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            if any(cell.value is not None for cell in row):
                return sheet
    raise InvalidWorkbookError("В книге нет ни одного непустого листа.")


def _read_header(sheet: Worksheet) -> tuple[int, dict[str, int]]:
    header_row_index = None
    for row in sheet.iter_rows():
        if any(cell.value is not None for cell in row):
            header_row_index = row[0].row
            break
    if header_row_index is None:
        raise InvalidWorkbookError("Не найдена строка заголовков.")

    header_cells = sheet[header_row_index]
    header_by_text: dict[str, int] = {}
    for cell in header_cells:
        if cell.value is None:
            continue
        text = str(cell.value).strip()
        if text:
            header_by_text[text] = cell.column

    missing = [col for col in REQUIRED_COLUMNS if col not in header_by_text]
    if missing:
        raise ImportValidationError(
            "В файле отсутствуют обязательные колонки.",
            details=[
                {
                    "row": header_row_index,
                    "field": col,
                    "code": "IMPORT_MISSING_COLUMN",
                    "message": f"Отсутствует обязательная колонка «{col}».",
                }
                for col in missing
            ],
        )

    columns = {col: header_by_text[col] for col in REQUIRED_COLUMNS}
    for col in OPTIONAL_COLUMNS:
        if col in header_by_text:
            columns[col] = header_by_text[col]

    return header_row_index, columns


def _row_is_empty(values: dict) -> bool:
    return all(cell.value is None or (isinstance(cell.value, str) and not cell.value.strip()) for cell in values.values())


def _is_formula_cell(cell) -> bool:
    return getattr(cell, "data_type", None) == "f"


def _issue(row: int, field: str | None, code: str, message: str) -> dict:
    return {"row": row, "field": field, "code": code, "message": message}


def _parse_name(cell, row: int, errors: list[dict]) -> str | None:
    if _is_formula_cell(cell):
        errors.append(_issue(row, "Задача", "IMPORT_FORMULA_NOT_SUPPORTED", "Формулы в поле «Задача» не поддерживаются."))
        return None
    value = cell.value
    text = str(value).strip() if value is not None else ""
    if not text:
        errors.append(_issue(row, "Задача", "IMPORT_MISSING_TASK_NAME", "Название задачи обязательно."))
        return None
    if len(text) > MAX_TASK_NAME_LENGTH:
        errors.append(
            _issue(row, "Задача", "IMPORT_INVALID_TASK_NAME", f"Название длиннее {MAX_TASK_NAME_LENGTH} символов.")
        )
        return None
    if ";" in text:
        errors.append(_issue(row, "Задача", "IMPORT_INVALID_TASK_NAME", "Название не может содержать «;»."))
        return None
    return text


def _parse_description(cell, row: int, errors: list[dict]) -> str:
    value = cell.value
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) > MAX_DESCRIPTION_LENGTH:
        errors.append(
            _issue(row, "Описание", "IMPORT_INVALID_DESCRIPTION", f"Описание длиннее {MAX_DESCRIPTION_LENGTH} символов.")
        )
    return text


def _parse_assignee(cell, row: int, errors: list[dict]) -> str | None:
    value = cell.value
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if len(text) > MAX_ASSIGNEE_LENGTH:
        errors.append(
            _issue(row, "Исполнитель", "IMPORT_INVALID_ASSIGNEE", f"Исполнитель длиннее {MAX_ASSIGNEE_LENGTH} символов.")
        )
        return None
    return text


def _parse_duration(cell, row: int, errors: list[dict]) -> int | None:
    if _is_formula_cell(cell):
        errors.append(
            _issue(row, "Длительность", "IMPORT_FORMULA_NOT_SUPPORTED", "Формулы в поле «Длительность» не поддерживаются.")
        )
        return None

    value = cell.value
    candidate: int | None = None

    if isinstance(value, bool):
        candidate = None
    elif isinstance(value, int):
        candidate = value
    elif isinstance(value, float):
        candidate = int(value) if value.is_integer() else None
    elif isinstance(value, str):
        stripped = value.strip()
        try:
            as_float = float(stripped)
        except ValueError:
            candidate = None
        else:
            candidate = int(as_float) if as_float.is_integer() else None

    if candidate is None or not (MIN_DURATION_WORKDAYS <= candidate <= MAX_DURATION_WORKDAYS):
        errors.append(
            _issue(
                row,
                "Длительность",
                "IMPORT_INVALID_DURATION",
                f"Длительность должна быть целым числом от {MIN_DURATION_WORKDAYS} до {MAX_DURATION_WORKDAYS}.",
            )
        )
        return None
    return candidate


def _parse_predecessors(cell, row: int, errors: list[dict]) -> list[str]:
    if _is_formula_cell(cell):
        errors.append(
            _issue(
                row,
                "Предшественники",
                "IMPORT_FORMULA_NOT_SUPPORTED",
                "Формулы в поле «Предшественники» не поддерживаются.",
            )
        )
        return []
    value = cell.value
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    return [part.strip() for part in text.split(";") if part.strip()]


def _parse_optional_date(cell, row: int, field: str, errors: list[dict]) -> date | None:
    if _is_formula_cell(cell):
        errors.append(
            _issue(row, field, "IMPORT_FORMULA_NOT_SUPPORTED", f"Формулы в поле «{field}» не поддерживаются.")
        )
        return None
    value = cell.value
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            return date.fromisoformat(stripped)
        except ValueError:
            errors.append(_issue(row, field, "IMPORT_INVALID_DATE", f"Не удалось разобрать дату «{stripped}»."))
            return None
    errors.append(_issue(row, field, "IMPORT_INVALID_DATE", "Неподдерживаемый формат даты."))
    return None
