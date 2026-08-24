"""Excel export (product-spec §19, integration brief EXCEL EXPORT)."""

from __future__ import annotations

import io
import uuid
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.models import Project, Task

_DATE_FORMAT = "DD.MM.YYYY"
_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@")

_PLAN_HEADERS = (
    "Задача",
    "Описание",
    "Исполнитель",
    "Длительность",
    "Предшественники",
    "Плановая трудоёмкость, ч",
    "Дата начала",
    "Дата окончания",
    "Не ранее",
)


class ExcelExportService:
    def export(self, project: Project) -> bytes:
        workbook = Workbook()
        plan_sheet = workbook.active
        plan_sheet.title = "План"
        self._write_plan_sheet(plan_sheet, project)

        meta_sheet = workbook.create_sheet("Метаданные")
        self._write_metadata_sheet(meta_sheet, project)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _write_plan_sheet(self, sheet: Worksheet, project: Project) -> None:
        sheet.append(_PLAN_HEADERS)

        by_id: dict[uuid.UUID, Task] = {t.id: t for t in project.tasks}
        ordered = sorted(project.tasks, key=lambda t: t.display_order)

        for task in ordered:
            predecessor_names = ";".join(
                _safe_text(by_id[p].name) for p in task.predecessor_ids if p in by_id
            )
            row_index = sheet.max_row + 1
            sheet.append(
                [
                    _safe_text(task.name),
                    _safe_text(task.description),
                    _safe_text(task.assignee) if task.assignee else None,
                    task.duration_workdays,
                    predecessor_names or None,
                    task.planned_effort_hours,
                    task.start_date,
                    task.end_date,
                    task.start_not_before,
                ]
            )
            for col in (7, 8, 9):
                cell = sheet.cell(row=row_index, column=col)
                if cell.value is not None:
                    cell.number_format = _DATE_FORMAT

    def _write_metadata_sheet(self, sheet: Worksheet, project: Project) -> None:
        sheet.append(["project_name", project.name])
        sheet.append(["project_start_date", project.project_start_date.isoformat()])
        sheet.append(["revision", project.revision])
        sheet.append(["exported_at_utc", datetime.now(timezone.utc).isoformat()])


def _safe_text(value: str) -> str:
    """Prefix a leading `=`, `+`, `-` or `@` with an apostrophe-equivalent guard
    so Excel/LibreOffice never interprets an imported string as a formula."""
    if value and value[0] in _FORMULA_TRIGGER_CHARS:
        return f"'{value}"
    return value


def export_filename(project: Project) -> str:
    safe_name = "".join(ch if ch.isalnum() or ch in "-_ " else "_" for ch in project.name).strip() or "project"
    return f"{safe_name}.xlsx"
