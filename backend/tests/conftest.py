from __future__ import annotations

import io
import json
import uuid
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook

from app.agent.conversation_store import InMemoryAgentConversationStore
from app.agent.openrouter_client import CompletionResult
from app.agent.service import AgentService
from app.domain.models import Project, Task
from app.mcp_server.app import build_mcp_asgi_app, create_mcp_server
from app.services.project_service import ProjectService
from app.settings import Settings
from app.storage.project_store import InMemoryProjectStore

REPO_ROOT = Path(__file__).resolve().parents[2]
SAMPLE_XLSX_PATH = REPO_ROOT / "examples" / "sample_patient_card_project.xlsx"
CANONICAL_START_DATE = date(2026, 9, 7)


def make_task(
    *,
    name: str,
    duration_workdays: int = 1,
    predecessor_ids: list[uuid.UUID] | None = None,
    display_order: int = 0,
    start_not_before: date | None = None,
    assignee: str | None = None,
    description: str = "",
    created_source: str = "import",
    task_id: uuid.UUID | None = None,
) -> Task:
    return Task(
        id=task_id or uuid.uuid4(),
        name=name,
        description=description,
        assignee=assignee,
        duration_workdays=duration_workdays,
        predecessor_ids=predecessor_ids or [],
        start_not_before=start_not_before,
        start_date=date(2000, 1, 1),
        end_date=date(2000, 1, 1),
        display_order=display_order,
        created_source=created_source,  # type: ignore[arg-type]
    )


def make_project(tasks: list[Task], *, revision: int = 1, project_start_date: date = CANONICAL_START_DATE) -> Project:
    now = datetime.now(timezone.utc)
    return Project(
        id=uuid.uuid4(),
        name="test-project",
        project_start_date=project_start_date,
        revision=revision,
        tasks=tasks,
        created_at=now,
        updated_at=now,
    )


@pytest.fixture
def sample_xlsx_path() -> Path:
    assert SAMPLE_XLSX_PATH.exists(), f"canonical sample workbook missing at {SAMPLE_XLSX_PATH}"
    return SAMPLE_XLSX_PATH


DEFAULT_HEADERS = ("Задача", "Описание", "Исполнитель", "Длительность", "Предшественники")


def build_workbook_bytes(
    rows: list[list],
    *,
    headers: tuple[str, ...] = DEFAULT_HEADERS,
    sheet_name: str = "План",
    formula_cells: dict[tuple[int, int], str] | None = None,
    merge_range: str | None = None,
    leading_blank_sheet: bool = False,
) -> bytes:
    """Build a minimal .xlsx in memory for import tests.

    `formula_cells` maps (row_index_1_based_in_data, col_index_1_based) to a
    formula string, e.g. {(1, 1): "=1+1"} sets a formula in the first data
    row's first column.
    """
    workbook = Workbook()

    if leading_blank_sheet:
        workbook.active.title = "Пусто"
        sheet = workbook.create_sheet(sheet_name)
    else:
        sheet = workbook.active
        sheet.title = sheet_name

    sheet.append(list(headers))
    for row in rows:
        sheet.append(row)

    if formula_cells:
        for (row_idx, col_idx), formula in formula_cells.items():
            cell = sheet.cell(row=row_idx + 1, column=col_idx)  # +1 for header row
            cell.value = formula

    if merge_range:
        sheet.merge_cells(merge_range)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# --- MCP / agent test helpers -----------------------------------------------------


@dataclass
class McpTestEnv:
    store: InMemoryProjectStore
    service: ProjectService
    mcp: Any
    asgi_app: Any


def build_mcp_test_env() -> McpTestEnv:
    store = InMemoryProjectStore()
    service = ProjectService(store)
    mcp = create_mcp_server(service)
    asgi_app = build_mcp_asgi_app(mcp)
    return McpTestEnv(store=store, service=service, mcp=mcp, asgi_app=asgi_app)


async def import_sample_via_service(service: ProjectService) -> Project:
    content = SAMPLE_XLSX_PATH.read_bytes()
    project, _warnings = await service.import_project(
        content=content, filename=SAMPLE_XLSX_PATH.name, project_start_date=CANONICAL_START_DATE
    )
    return project


def build_test_settings(**overrides: Any) -> Settings:
    """The one place tests construct `Settings`. Always isolated from the
    developer's local `backend/.env` (`_env_file=None`) so a value tuned
    there for manual/live runs — e.g. a higher AGENT_MAX_STEPS for a
    reasoning-heavy free model — can never change what a unit/integration
    test asserts. Every test-side `Settings(...)` call should go through
    this factory instead of constructing `Settings` directly."""
    return Settings(_env_file=None, **overrides)


def build_agent_service(env: McpTestEnv, fake_client: "FakeOpenRouterClient", **settings_overrides: Any) -> AgentService:
    settings = build_test_settings(
        openrouter_api_key="test-key",
        openrouter_model="test-model",
        **settings_overrides,
    )
    return AgentService(
        settings=settings,
        project_service=env.service,
        mcp_asgi_app=env.asgi_app,
        conversation_store=InMemoryAgentConversationStore(max_turns=settings.agent_history_turns),
        openrouter_client=fake_client,
    )


ScriptedToolCall = dict[str, Any]  # {"name": str, "arguments": dict}


class FakeOpenRouterClient:
    """A scripted stand-in for OpenRouterClient — no network. Each item in
    `script` is either a final-text string, a list of tool calls (dicts with
    "name"/"arguments") to return as one assistant message, or an Exception
    instance to raise on that turn."""

    def __init__(self, script: list[str | list[ScriptedToolCall] | Exception]) -> None:
        self._script = list(script)
        self.calls: list[dict[str, Any]] = []

    async def chat_completion(
        self, *, messages: list[dict[str, Any]], tools: list[dict[str, Any]] | None = None
    ) -> CompletionResult:
        self.calls.append({"messages": [dict(m) for m in messages], "tools": tools})
        if not self._script:
            raise AssertionError("FakeOpenRouterClient script exhausted — test scripted too few turns")
        item = self._script.pop(0)

        if isinstance(item, Exception):
            raise item

        if isinstance(item, str):
            return CompletionResult(
                message={"role": "assistant", "content": item}, finish_reason="stop", usage=None, latency_ms=1.0
            )

        tool_calls = [
            {
                "id": f"call_{i}",
                "type": "function",
                "function": {"name": call["name"], "arguments": json.dumps(call["arguments"], ensure_ascii=False)},
            }
            for i, call in enumerate(item)
        ]
        return CompletionResult(
            message={"role": "assistant", "content": None, "tool_calls": tool_calls},
            finish_reason="tool_calls",
            usage={"total_tokens": 10},
            latency_ms=1.0,
        )
