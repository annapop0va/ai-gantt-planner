from __future__ import annotations

import io
from datetime import date

import openpyxl

from app.services.excel_export import ExcelExportService
from app.services.excel_import import ExcelImportService
from tests.conftest import make_project, make_task


def test_export_canonical_18_tasks_round_trips():
    a = make_task(name="A", duration_workdays=2, display_order=1)
    b = make_task(name="B", duration_workdays=3, predecessor_ids=[a.id], display_order=2, assignee="Alice")
    tasks = [a, b]
    project = make_project(tasks)

    content = ExcelExportService().export(project)
    result = ExcelImportService().parse(content, filename="roundtrip.xlsx")
    assert len(result.tasks) == 2
    by_name = {t.name: t for t in result.tasks}
    assert by_name["B"].predecessor_ids == [by_name["A"].id]
    assert by_name["B"].assignee == "Alice"
    assert by_name["A"].assignee is None


def test_predecessors_exported_as_display_names():
    a = make_task(name="Requirements", duration_workdays=1, display_order=1)
    b = make_task(name="Build", duration_workdays=1, predecessor_ids=[a.id], display_order=2)
    project = make_project([a, b])

    content = ExcelExportService().export(project)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["План"]
    header = [c.value for c in ws[1]]
    pred_col = header.index("Предшественники") + 1
    assert ws.cell(row=3, column=pred_col).value == "Requirements"


def test_dates_are_real_date_cells_with_display_format():
    a = make_task(name="A", duration_workdays=1, display_order=1)
    project = make_project([a])
    from app.scheduler.engine import compute_schedule

    schedule = compute_schedule(project.tasks, project.project_start_date)
    project = project.model_copy(update={"tasks": schedule.tasks})

    content = ExcelExportService().export(project)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["План"]
    header = [c.value for c in ws[1]]
    start_col = header.index("Дата начала") + 1
    cell = ws.cell(row=2, column=start_col)
    assert isinstance(cell.value, date)
    assert cell.number_format == "DD.MM.YYYY"


def test_effort_is_recalculated_not_read_from_task():
    a = make_task(name="A", duration_workdays=4, display_order=1)
    project = make_project([a])
    content = ExcelExportService().export(project)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["План"]
    header = [c.value for c in ws[1]]
    effort_col = header.index("Плановая трудоёмкость, ч") + 1
    assert ws.cell(row=2, column=effort_col).value == 32


def test_metadata_sheet_contains_expected_fields():
    a = make_task(name="A", duration_workdays=1, display_order=1)
    project = make_project([a], revision=3)
    content = ExcelExportService().export(project)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["Метаданные"]
    rows = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)}
    assert rows["project_name"] == project.name
    assert rows["project_start_date"] == project.project_start_date.isoformat()
    assert rows["revision"] == 3
    assert "exported_at_utc" in rows


def test_formula_injection_is_neutralized():
    a = make_task(name="=cmd|'/c calc'!A1", duration_workdays=1, display_order=1, assignee="+1;evil")
    project = make_project([a])
    content = ExcelExportService().export(project)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["План"]
    name_cell = ws.cell(row=2, column=1)
    assignee_cell = ws.cell(row=2, column=3)
    assert name_cell.data_type != "f"
    assert str(name_cell.value).startswith("'=")
    assert str(assignee_cell.value).startswith("'+")


def test_empty_assignee_exports_empty_cell():
    a = make_task(name="A", duration_workdays=1, display_order=1, assignee=None)
    project = make_project([a])
    content = ExcelExportService().export(project)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["План"]
    header = [c.value for c in ws[1]]
    assignee_col = header.index("Исполнитель") + 1
    assert ws.cell(row=2, column=assignee_col).value is None
